# -*- coding: utf-8 -*-
# 导入load_workbook
from openpyxl import load_workbook

#读取测试数据
#将excel中每一条测试用例读取到一个列表中

#读取一条测试用例——写到一个函数中
def read_data(sheet_name,case_id):
 # 打开excel
 workbook1=load_workbook('test_case2.xlsx')
 # 定位表单（test_data）
 sheet1=workbook1[sheet_name]
 print(sheet1)
 test_case=[] #用来存储每一行数据，也就是一条测试用例
 test_case.append(sheet1.cell(case_id+1,1).value)
 test_case.append(sheet1.cell(case_id+1,2).value)
 test_case.append(sheet1.cell(case_id+1,3).value)
 test_case.append(sheet1.cell(case_id+1,4).value)
 test_case.append(sheet1.cell(case_id+1,5).value)
 test_case.append(sheet1.cell(case_id+1,6).value)
 return test_case #将读取到的用例返回

#调用函数读取第1条测试用例，并将返回结果保存在data中
# data=read_data(1)
# print(data)

#将测试结果写会excel
def write_data(sheet_name,row,col,value):
 workbook1=load_workbook('test_case2.xlsx')
 sheet=workbook1[sheet_name]
 sheet.cell(row,col).value=value
 workbook1.save('test_case2.xlsx')


#统计测试用例的行数
def count_case(sheet_name):
  workbook1=load_workbook('test_case2.xlsx')
  sheet=workbook1[sheet_name]
  max_row=sheet.max_row #统计测试用例的行数
  return max_row

